"""
Market Tracker v2 - Índices, Bonos, Commodities + Macro Dashboard
Requiere: pip install yfinance openpyxl
"""

import yfinance as yf
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# ── UNIVERSO DE ACTIVOS ───────────────────────────────────────────────────────

SECTIONS = [
    ("MAJOR INDICES", [
        ("S&P 500",               "^GSPC"),
        ("MSCI World",            "URTH"),
        ("NASDAQ Composite",      "^IXIC"),
        ("Euro STOXX 50",         "^STOXX50E"),
        ("MSCI Emerging Markets", "EEM"),
        ("Russell 2000",          "^RUT"),
        ("S&P 500 Equal Weight",  "RSP"),
    ]),
    ("MAG7", [
        ("NVIDIA",         "NVDA"),
        ("Apple",          "AAPL"),
        ("Alphabet",       "GOOGL"),
        ("Microsoft",      "MSFT"),
        ("Amazon",         "AMZN"),
        ("Meta Platforms", "META"),
        ("Tesla",          "TSLA"),
    ]),
    ("US TREASURIES (ETFs)", [
        ("TLT (20Y+ ETF)",  "TLT"),
        ("IEF (7-10Y ETF)", "IEF"),
        ("IEI (3-7Y ETF)",  "IEI"),
        ("SHY (1-3Y ETF)",  "SHY"),
        ("BIL (0-3M ETF)",  "BIL"),
    ]),
    ("GLOBAL BONDS (ETFs)", [
        ("Bund 10Y · iShares (EUR)", "EXX6.DE"),   # iShares Bund 7-10Y
        ("Gilt 10Y · iShares (GBP)", "IGLT.L"),    # iShares Core UK Gilts
        ("JGB · iShares (JPY)",      "2561.T"),     # iShares JPY Govt Bond
        ("Italy BTP · iShares",      "ITPS.MI"),   # iShares BTP
        ("EU Govt Bond · iShares",   "IEAG.AS"),   # iShares Core Euro Govt
        ("EM Bond · iShares (USD)",  "EMB"),        # iShares EM USD Bond
    ]),
    ("COMMODITIES", [
        ("Gold",        "GC=F"),
        ("Silver",      "SI=F"),
        ("Oil (WTI)",   "CL=F"),
        ("Oil (Brent)", "BZ=F"),
        ("Natural Gas", "NG=F"),
        ("Copper",      "HG=F"),
        ("Bitcoin",     "BTC-USD"),
        ("Ethereum",    "ETH-USD"),
    ]),
    ("EUROPE", [
        ("United Kingdom (FTSE 100)", "^FTSE"),
        ("France (CAC 40)",           "^FCHI"),
        ("Germany (DAX)",             "^GDAXI"),
        ("Netherlands (AEX)",         "^AEX"),
        ("Spain (IBEX 35)",           "^IBEX"),
    ]),
    ("ASIA", [
        ("Japan",      "^N225"),
        ("South Korea","^KS11"),
        ("India",      "^BSESN"),
        ("China",      "MCHI"),
        ("Hong Kong",  "^HSI"),
    ]),
    ("LATAM", [
        ("Brazil",    "EWZ"),
        ("Mexico",    "EWW"),
        ("Argentina", "ARGT"),
    ]),
    ("US SECTORS", [
        ("Technology",             "XLK"),
        ("Healthcare",             "XLV"),
        ("Financials",             "XLF"),
        ("Consumer Discretionary", "XLY"),
        ("Communication Services", "XLC"),
        ("Industrials",            "XLI"),
        ("Consumer Staples",       "XLP"),
        ("Energy",                 "XLE"),
        ("Utilities",              "XLU"),
        ("Real Estate",            "XLRE"),
        ("Materials",              "XLB"),
    ]),
    ("EU SECTORS", [
        ("EU Banks",                  "EXV1.DE"),
        ("EU Healthcare",             "EXH1.DE"),
        ("EU Industrials",            "EXH2.DE"),
        ("EU Energy",                 "EXH8.DE"),
        ("EU Technology",             "EXV4.DE"),
        ("EU Consumer Discretionary", "EXH3.DE"),
        ("EU Consumer Staples",       "EXH4.DE"),
        ("EU Telecom",                "EXH6.DE"),
        ("EU Utilities",              "EXH7.DE"),
        ("EU Materials",              "EXH5.DE"),
        ("EU Real Estate",            "IPRP.AS"),
    ]),
]

# ── MACRO ─────────────────────────────────────────────────────────────────────

MACRO_DATA = {
    "US YIELD CURVE": [
        ("US 3M Yield",  "DGS3MO", "%", "FRED", "T-Bill 3 meses"),
        ("US 2Y Yield",  "DGS2",   "%", "FRED", "Treasury 2 años"),
        ("US 5Y Yield",  "DGS5",   "%", "FRED", "Treasury 5 años"),
        ("US 10Y Yield", "DGS10",  "%", "FRED", "Treasury 10 años"),
        ("US 30Y Yield", "DGS30",  "%", "FRED", "Treasury 30 años"),
        ("10Y - 2Y",     "T10Y2Y", "pts", "FRED", "Spread curva. Negativo = invertida"),
        ("10Y - 3M",     "T10Y3M", "pts", "FRED", "Spread curva corta"),
    ],
    "UNITED STATES": [
        ("Fed Funds Rate",    "FEDFUNDS",              "%",   "FRED", "Target upper bound"),
        ("CPI YoY",           "CPIAUCSL",              "%",   "FRED", "All items, not seas. adj."),
        ("Core CPI YoY",      "CPILFESL",              "%",   "FRED", "Ex food & energy"),
        ("GDP Growth QoQ",    "A191RL1Q225SBEA",       "%",   "FRED", "Real GDP, annualised"),
        ("Unemployment Rate", "UNRATE",                "%",   "FRED", "U-3 rate"),
        ("10Y-2Y Spread",     "T10Y2Y",                "pts", "FRED", "Negative = inverted curve"),
        ("VIX",               "^VIX",                  "pts", "YF",   ">30 = high fear"),
        ("DXY (USD Index)",   "DX-Y.NYB",              "pts", "YF",   "Trade-weighted USD"),
        ("OECD CLI",          "USALOLITONOSTSAM",      "pts", "FRED", "CLI >100 expansión, <100 contracción"),
        ("CFNAI",             "CFNAI",                 "pts", "FRED", "Chicago Fed. <-0.7 = recesión"),
    ],
    "EURO ZONE": [
        ("ECB Deposit Rate",  "ECBDFR",                "%",   "FRED", "ECB deposit facility rate"),
        ("CPI YoY (EA)",      "CP0000EZ19M086NEST",    "%",   "FRED", "Euro Area HICP"),
        ("Unemployment (EA)", "LRHUTTTTEZM156S",       "%",   "FRED", "Euro Area unemployment"),
        ("GDP Growth QoQ",    "CLVMNACSCAB1GQEA19",    "%",   "FRED", "Euro Area real GDP"),
        ("EUR/USD",           "EURUSD=X",              "fx",  "YF",   "Spot rate"),
        ("GBP/USD",           "GBPUSD=X",              "fx",  "YF",   "Spot rate"),
        ("USD/JPY",           "JPY=X",                 "fx",  "YF",   "JPY per USD"),
    ],
}

# ── CÁLCULOS ──────────────────────────────────────────────────────────────────

def get_prices(tickers):
    end   = datetime.today()
    start = end - timedelta(days=365 * 5 + 10)
    raw   = yf.download(tickers, start=start, end=end, auto_adjust=True, progress=False)
    return raw["Close"]

def calc_return(series, days):
    s = series.dropna()
    if len(s) < 2: return None
    past = s.iloc[-days] if len(s) > days else s.iloc[0]
    return (s.iloc[-1] - past) / past if past != 0 else None

def ytd_return(series):
    s = series.dropna()
    if len(s) < 2: return None
    ytd_s = s[s.index >= datetime(datetime.today().year, 1, 1)]
    if len(ytd_s) < 1: return None
    return (s.iloc[-1] - ytd_s.iloc[0]) / ytd_s.iloc[0] if ytd_s.iloc[0] != 0 else None

def pct_from_low_high(series):
    s = series.dropna().iloc[-252:]
    if len(s) == 0: return None, None
    p = series.dropna().iloc[-1]
    lo, hi = s.min(), s.max()
    return ((p-lo)/lo if lo!=0 else None, (p-hi)/hi if hi!=0 else None)

def build_market_rows(prices_df):
    rows = []
    for section_name, assets in SECTIONS:
        rows.append(("HEADER", section_name))
        for name, ticker in assets:
            if ticker not in prices_df.columns:
                rows.append((name, ticker, None, None, None, None, None, None, None, None, None))
                continue
            s = prices_df[ticker]
            price = float(s.dropna().iloc[-1]) if len(s.dropna()) > 0 else None
            pct_low, pct_high = pct_from_low_high(s)
            rows.append((name, ticker, price, pct_low, pct_high,
                         calc_return(s,1), calc_return(s,5), calc_return(s,21),
                         ytd_return(s), calc_return(s,252), calc_return(s,252*5)))
    return rows

# ── Caché FRED — persiste últimos valores buenos entre ejecuciones ────────────
FRED_CACHE_FILE = "fred_cache.json"

def load_fred_cache():
    import json, os
    if os.path.exists(FRED_CACHE_FILE):
        try:
            with open(FRED_CACHE_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_fred_cache(cache):
    import json
    try:
        with open(FRED_CACHE_FILE, "w") as f:
            json.dump(cache, f)
    except Exception:
        pass

_fred_cache = None

def get_fred_series(series_id):
    """Descarga datos de FRED via API oficial con API key. Si falla, usa caché."""
    global _fred_cache
    import urllib.request, json, os

    if _fred_cache is None:
        _fred_cache = load_fred_cache()

    api_key = os.environ.get("FRED_API_KEY", "")
    try:
        if api_key:
            url = (f"https://api.stlouisfed.org/fred/series/observations"
                   f"?series_id={series_id}&api_key={api_key}&file_type=json"
                   f"&sort_order=desc&limit=20")
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=15) as r:
                data = json.loads(r.read().decode())
            obs = [o for o in data.get("observations", []) if o["value"] not in (".", "")]
            if not obs:
                raise ValueError("No observations")
            vals = [float(o["value"]) for o in reversed(obs)]
        else:
            url = f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}"
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=15) as r:
                lines = r.read().decode().strip().split("\n")[1:]
            vals = []
            for line in lines:
                parts = line.split(",")
                if len(parts) == 2 and parts[1].strip() not in ("", "."):
                    try: vals.append(float(parts[1]))
                    except ValueError: pass
            if not vals:
                raise ValueError("No data")

        latest   = vals[-1]
        chg_last = vals[-1] - vals[-2]  if len(vals) > 1  else None
        chg_yoy  = vals[-1] - vals[-13] if len(vals) > 13 else None
        # Guardar en caché
        _fred_cache[series_id] = [latest, chg_last, chg_yoy]
        save_fred_cache(_fred_cache)
        return latest, chg_last, chg_yoy

    except Exception as e:
        # Intentar caché
        if series_id in _fred_cache:
            cached = _fred_cache[series_id]
            print(f"     ⚠️  FRED {series_id} falla ({e}) — usando caché: {cached[0]}")
            return tuple(cached)
        print(f"     ⚠️  FRED {series_id}: {e} — sin caché disponible")
        return None, None, None

# ── Historial SPI diario ──────────────────────────────────────────────────────
SPI_HISTORY_FILE = "spi_history.json"

def save_spi_history(today_str, degrees, phase_idx, signals):
    """Guarda los datos del día en el historial SPI."""
    import json, os

    def sv(name):
        s = signals.get(name, (None, None, None))
        return round(s[1], 4) if s[1] is not None else None

    record = {
        "degrees":  round(degrees, 2),
        "phase":    PHASE_NAMES[phase_idx],
        "gdp":      sv("GDP QoQ"),
        "unemp":    sv("Desempleo"),
        "cpi":      sv("CPI YoY"),
        "fed":      sv("Fed Funds"),
        "curva":    sv("Curva 10Y-2Y"),
        "y10":      sv("10Y Yield"),
        "cli":      sv("OECD CLI"),
        "cfnai":    sv("CFNAI"),
        "vix_ma25": sv("VIX MA25/200"),
    }

    # Cargar historial existente
    history = {}
    if os.path.exists(SPI_HISTORY_FILE):
        try:
            with open(SPI_HISTORY_FILE) as f:
                history = json.load(f)
        except Exception:
            history = {}

    history[today_str] = record

    # Guardar ordenado por fecha
    history = dict(sorted(history.items(), reverse=True))
    try:
        with open(SPI_HISTORY_FILE, "w") as f:
            json.dump(history, f, indent=2, ensure_ascii=False)
        print(f"   📅 Historial SPI guardado ({len(history)} días)")
    except Exception as e:
        print(f"   ⚠️  Error guardando historial: {e}")


def get_yf_macro(ticker):
    try:
        df = yf.download(ticker, period="2y", auto_adjust=True, progress=False)
        if df.empty: return None, None, None
        s = df["Close"].dropna()
        if hasattr(s, 'columns'):  # MultiIndex — coger primera columna
            s = s.iloc[:, 0]
        latest   = float(s.iloc[-1].item() if hasattr(s.iloc[-1], 'item') else s.iloc[-1])
        chg_last = float((s.iloc[-1] - s.iloc[-2]).item()) if len(s) > 1 else None
        chg_yoy  = float((s.iloc[-1] - s.iloc[-252]).item()) if len(s) > 252 else None
        return latest, chg_last, chg_yoy
    except Exception:
        return None, None, None

# Series FRED que son índices de nivel (no % directos) — hay que calcular variación YoY
FRED_INDEX_SERIES = {"CPIAUCSL", "CPILFESL", "CP0000EZ19M086NEST", "CLVMNACSCAB1GQEA19"}

def build_macro_data():
    results = {}
    for region, indicators in MACRO_DATA.items():
        results[region] = []
        for name, series_id, unit, source, note in indicators:
            print(f"  📊 {name}...")
            if source == "YF":
                val, chg_last, chg_yoy = get_yf_macro(series_id)
            else:
                val, chg_last, chg_yoy = get_fred_series(series_id)
                # Convertir índices de nivel a variación YoY%
                if series_id in FRED_INDEX_SERIES and val is not None and chg_yoy is not None:
                    base = val - chg_yoy
                    if base != 0:
                        val      = (chg_yoy / base) * 100   # YoY%
                        chg_last = None                      # chg_last en puntos de índice no es útil
                        chg_yoy  = None                      # ya está incorporado en val
            results[region].append((name, unit, val, chg_last, chg_yoy, note))
    return results

# ── ESTILOS ───────────────────────────────────────────────────────────────────

DARK_BG    = "1C2634"
HEADER_BG  = "2C3E50"
SECTION_BG = "243447"
ACCENT     = "1A6B9A"
WHITE      = "FFFFFF"
LIGHT_GRAY = "D6DCE4"
MID_GRAY   = "8496A9"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def fnt(bold=False, color=WHITE, size=9, italic=False):
    return Font(bold=bold, color=color, size=size, name="Arial", italic=italic)

def center():
    return Alignment(horizontal="center", vertical="center")

def left(indent=0):
    return Alignment(horizontal="left", vertical="center", indent=indent)

def color_ret(val):
    if val is None:        return HEADER_BG
    if val >=  0.10:       return "1A7A3A"
    if val >=  0.05:       return "27AE60"
    if val >=  0.02:       return "52BE80"
    if val >=  0.00:       return "A9DFBF"
    if val >= -0.02:       return "F1948A"
    if val >= -0.05:       return "E74C3C"
    if val >= -0.10:       return "C0392B"
    return "922B21"

def text_ret(val):
    if val is None: return MID_GRAY
    return "2C2C2C" if abs(val) < 0.02 else WHITE

def color_chg(val):
    if val is None or val == 0: return SECTION_BG
    return "27AE60" if val > 0 else "E74C3C"

def fmt_pct(val):
    if val is None: return "–"
    return f"{'+'if val>0 else ''}{val*100:.1f}%"

def fmt_price(val):
    if val is None: return "–"
    if val >= 10000: return f"{val:,.0f}"
    if val >= 100:   return f"{val:,.1f}"
    if val >= 1:     return f"{val:,.2f}"
    return f"{val:.4f}"

def fmt_macro(val, unit):
    if val is None: return "N/A"
    if unit == "%":   return f"{val:.2f}%"
    if unit == "pts": return f"{val:.2f}"
    if unit == "fx":  return f"{val:.4f}"
    return f"{val:.2f}"

def fmt_chg(val, unit):
    if val is None: return "–"
    sign = "+" if val > 0 else ""
    if unit == "%":   return f"{sign}{val:.2f}pp"
    if unit == "pts": return f"{sign}{val:.2f}"
    if unit == "fx":  return f"{sign}{val:.4f}"
    return f"{sign}{val:.2f}"

# ── SHEET 1: MARKETS ──────────────────────────────────────────────────────────

MKT_HEADERS = ["Name", "Ticker", "Price", "vs 52W Low", "vs 52W High",
               "1D", "1W", "1M", "YTD", "1Y", "5Y"]
MKT_WIDTHS  = [28, 9, 11, 11, 11, 8, 8, 8, 8, 8, 8]


def add_banner_to_sheet(ws, banner_path, last_row):
    """Añade el banner de Quantfury al pie de la hoja."""
    import os
    if not os.path.exists(banner_path):
        return
    try:
        img = XLImage(banner_path)
        # Ancho total de tabla (~15 columnas × 60px ≈ 900px)
        img.width  = 900
        img.height = int(156 * 900 / 1919)
        # Colocar inmediatamente después del último contenido sin fila de separación
        cell = f"A{last_row + 1}"
        ws.add_image(img, cell)
    except Exception as e:
        print(f"   ⚠️  Banner no añadido: {e}")

def write_market_sheet(ws, rows, today_str):
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "Market Tracker"
    ws["A1"].font = fnt(bold=True, size=14); ws["A1"].fill = fill(DARK_BG); ws["A1"].alignment = left(1)

    ws.merge_cells("I1:K1")
    ws["I1"] = today_str
    ws["I1"].font = fnt(bold=True, color=LIGHT_GRAY, size=11); ws["I1"].fill = fill(DARK_BG); ws["I1"].alignment = center()

    ws.merge_cells("A2:K2")
    ws["A2"] = "All returns in local currency  |  Bond tickers show yield level or ETF price"
    ws["A2"].font = fnt(italic=True, color=MID_GRAY, size=8); ws["A2"].fill = fill(DARK_BG); ws["A2"].alignment = left(1)

    for rng, label in [("A3:B3","Overview"),("C3:E3","52W Range"),("F3:K3","Returns")]:
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = label; c.font = fnt(bold=True, color=LIGHT_GRAY, size=8)
        c.fill = fill(HEADER_BG); c.alignment = center()

    for ci, (h, w) in enumerate(zip(MKT_HEADERS, MKT_WIDTHS), 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = fnt(bold=True, color=LIGHT_GRAY, size=8); c.fill = fill(HEADER_BG); c.alignment = center()
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[3].height = 13
    ws.row_dimensions[4].height = 15

    er = 5
    for rd in rows:
        if rd[0] == "HEADER":
            ws.merge_cells(f"A{er}:K{er}")
            c = ws.cell(row=er, column=1, value=rd[1])
            c.font = fnt(bold=True, color=LIGHT_GRAY, size=8); c.fill = fill(SECTION_BG); c.alignment = left(1)
            ws.row_dimensions[er].height = 13; er += 1; continue

        name, ticker, price, pct_low, pct_high, r1d, r1w, r1m, rytd, r1y, r5y = rd
        vals = [name, ticker, fmt_price(price), fmt_pct(pct_low), fmt_pct(pct_high),
                fmt_pct(r1d), fmt_pct(r1w), fmt_pct(r1m), fmt_pct(rytd), fmt_pct(r1y), fmt_pct(r5y)]
        rets = [None,None,None,None,None, r1d, r1w, r1m, rytd, r1y, r5y]

        for ci, (v, r) in enumerate(zip(vals, rets), 1):
            c = ws.cell(row=er, column=ci, value=v)
            c.alignment = left(1) if ci <= 2 else center()
            if   ci == 1: c.fill=fill(DARK_BG);    c.font=fnt(color=LIGHT_GRAY, size=8)
            elif ci == 2: c.fill=fill(DARK_BG);    c.font=fnt(color=MID_GRAY, size=8, italic=True)
            elif ci == 3: c.fill=fill(DARK_BG);    c.font=fnt(size=8)
            elif ci in (4,5): c.fill=fill(SECTION_BG); c.font=fnt(color=LIGHT_GRAY, size=8)
            else:
                c.fill = fill(color_ret(r))
                c.font = fnt(color=text_ret(r), size=8, bold=(r is not None and abs(r)>0.05))
        ws.row_dimensions[er].height = 14; er += 1

    ws.freeze_panes = "A5"

# ── SHEET 2: MACRO ────────────────────────────────────────────────────────────

MAC_HEADERS = ["Indicator", "Unit", "Latest", "Chg (last period)", "Chg (YoY/12m)", "Note"]
MAC_WIDTHS  = [30, 7, 13, 18, 15, 38]

def write_macro_sheet(ws, macro_data, today_str):
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"] = "Macro Dashboard"
    ws["A1"].font = fnt(bold=True, size=14); ws["A1"].fill = fill(DARK_BG); ws["A1"].alignment = left(1)

    ws["F1"] = today_str
    ws["F1"].font = fnt(bold=True, color=LIGHT_GRAY, size=11); ws["F1"].fill = fill(DARK_BG); ws["F1"].alignment = center()

    ws.merge_cells("A2:F2")
    ws["A2"] = "FRED series via pandas_datareader  |  YF = Yahoo Finance  |  Data may lag 1-2 months for official stats"
    ws["A2"].font = fnt(italic=True, color=MID_GRAY, size=8); ws["A2"].fill = fill(DARK_BG); ws["A2"].alignment = left(1)

    for ci, (h, w) in enumerate(zip(MAC_HEADERS, MAC_WIDTHS), 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 13

    er = 3
    for region, indicators in macro_data.items():
        ws.merge_cells(f"A{er}:F{er}")
        c = ws.cell(row=er, column=1, value=region)
        c.font = fnt(bold=True, color=LIGHT_GRAY, size=9); c.fill = fill(SECTION_BG); c.alignment = left(1)
        ws.row_dimensions[er].height = 15; er += 1

        for ci, h in enumerate(MAC_HEADERS, 1):
            c = ws.cell(row=er, column=ci, value=h)
            c.font = fnt(bold=True, color=LIGHT_GRAY, size=8); c.fill = fill(HEADER_BG); c.alignment = center()
        ws.row_dimensions[er].height = 14; er += 1

        for name, unit, val, chg_last, chg_yoy, note in indicators:
            row_vals = [name, unit, fmt_macro(val, unit), fmt_chg(chg_last, unit), fmt_chg(chg_yoy, unit), note]
            for ci, rv in enumerate(row_vals, 1):
                c = ws.cell(row=er, column=ci, value=rv)
                c.alignment = left(1) if ci in (1, 6) else center()
                if   ci == 1: c.fill=fill(DARK_BG);    c.font=fnt(color=LIGHT_GRAY, size=8)
                elif ci == 2: c.fill=fill(DARK_BG);    c.font=fnt(color=MID_GRAY, size=8)
                elif ci == 3: c.fill=fill(DARK_BG);    c.font=fnt(bold=True, size=9)
                elif ci == 4:
                    bg = color_chg(chg_last)
                    c.fill=fill(bg); c.font=fnt(color=WHITE if bg!=SECTION_BG else MID_GRAY, size=8)
                elif ci == 5:
                    bg = color_chg(chg_yoy)
                    c.fill=fill(bg); c.font=fnt(color=WHITE if bg!=SECTION_BG else MID_GRAY, size=8)
                else: c.fill=fill(DARK_BG); c.font=fnt(color=MID_GRAY, size=8, italic=True)
            ws.row_dimensions[er].height = 14; er += 1

        er += 1

    ws.freeze_panes = "A3"

# ── SPI: SECTOR PULSE INVESTING ───────────────────────────────────────────────

# Pesos por fase (Expansión Temprana, Expansión Tardía, Recesión Temprana, Recesión Tardía)
SPI_SECTORS = [
    # (Nombre,           Ticker, ExpT,  ExpL,  RecT,  RecL)
    ("Technology",       "XLK",  0.25,  0.08,  0.05,  0.08),
    ("Cons. Discret.",   "XLY",  0.20,  0.08,  0.05,  0.05),
    ("Communication",    "XLC",  0.15,  0.07,  0.05,  0.05),
    ("Financials",       "XLF",  0.08,  0.20,  0.05,  0.05),
    ("Industrials",      "XLI",  0.08,  0.20,  0.05,  0.05),
    ("Materials",        "XLB",  0.05,  0.15,  0.05,  0.05),
    ("Cons. Staples",    "XLP",  0.05,  0.05,  0.20,  0.15),
    ("Healthcare",       "XLV",  0.05,  0.05,  0.25,  0.15),
    ("Utilities",        "XLU",  0.04,  0.04,  0.20,  0.10),
    ("Energy",           "XLE",  0.03,  0.08,  0.05,  0.15),
    ("Real Estate",      "XLRE", 0.02,  0.00,  0.00,  0.12),
]

PHASE_NAMES  = ["Expansión Temprana", "Expansión Tardía", "Recesión Temprana", "Recesión Tardía"]
PHASE_COLORS = ["27AE60", "F39C12", "E74C3C", "8E44AD"]  # verde, naranja, rojo, morado
PHASE_KEYS   = ["exp_early", "exp_late", "rec_early", "rec_late"]

# Indicadores macro usados para determinar la fase
# Cada indicador devuelve su último valor y su variación (chg)
# Lógica de scoring:
#   gdp_growing + unemp_falling + cpi_low  + rates_low/falling  → Expansión Temprana
#   gdp_growing + unemp_stable  + cpi_high + rates_rising        → Expansión Tardía
#   gdp_falling + unemp_rising  + cpi_high + rates_high          → Recesión Temprana
#   gdp_low     + unemp_high    + cpi_low  + rates_falling       → Recesión Tardía

def detect_cycle_phase(macro_results, prices_daily=None):
    """
    Sistema de promedio circular ponderado para determinar fase del ciclo.
    Cada indicador vota a un angulo objetivo con un peso.
    0-90: ExpT | 90-180: ExpL | 180-270: RecT | 270-360: RecL
    """
    import math

    EXP_T, EXP_L, REC_T, REC_L = 45, 135, 225, 315

    def weighted_circular_mean(votes):
        if not votes: return 135.0
        sin_sum = sum(w * math.sin(math.radians(a)) for a, w in votes)
        cos_sum = sum(w * math.cos(math.radians(a)) for a, w in votes)
        return math.degrees(math.atan2(sin_sum, cos_sum)) % 360

    def get_val(region, name):
        for r, indicators in macro_results.items():
            for n, unit, val, chg_last, chg_yoy, note in indicators:
                if n == name:
                    return val, chg_last
        return None, None

    def get_val_yoy(region, name):
        for r, indicators in macro_results.items():
            for n, unit, val, chg_last, chg_yoy, note in indicators:
                if n == name:
                    return val, chg_yoy
        return None, None

    gdp_val,    gdp_chg    = get_val("UNITED STATES", "GDP Growth QoQ")
    unemp_val,  unemp_chg  = get_val("UNITED STATES", "Unemployment Rate")
    cpi_val,    cpi_chg    = get_val("UNITED STATES", "CPI YoY")
    fed_val,    fed_chg    = get_val("UNITED STATES", "Fed Funds Rate")
    _, fed_yoy             = get_val_yoy("UNITED STATES", "Fed Funds Rate")
    spread_val, spread_chg = get_val("US YIELD CURVE", "10Y - 2Y")
    y10_val,    y10_chg    = get_val("US YIELD CURVE", "US 10Y Yield")
    cli_val,    cli_chg    = get_val("UNITED STATES", "OECD CLI")
    cfnai_val,  cfnai_chg  = get_val("UNITED STATES", "CFNAI")

    # ── Señales derivadas ─────────────────────────────────────────────────────
    gdp_growing    = gdp_val   is not None and gdp_val   > 0
    unemp_falling  = unemp_chg is not None and unemp_chg < 0
    unemp_rising   = unemp_chg is not None and unemp_chg > 0
    unemp_high     = unemp_val is not None and unemp_val > 5.0
    cpi_high       = cpi_val   is not None and cpi_val   > 3.0
    cpi_rising     = cpi_chg   is not None and cpi_chg   > 0
    rates_rising   = fed_yoy   is not None and fed_yoy   > 0.25
    rates_falling  = fed_yoy   is not None and fed_yoy   < -0.25
    rates_high     = fed_val   is not None and fed_val   > 4.0

    # Curva
    curve_inverted    = spread_val is not None and spread_val < 0
    curve_normalizing = (spread_val is not None and spread_chg is not None
                         and spread_val < 0.5 and spread_chg > 0)
    curve_normal      = spread_val is not None and 0.3 < spread_val <= 1.5
    curve_steep       = spread_val is not None and spread_val > 1.5

    if spread_val is not None:
        if curve_inverted:      curve_txt = "Invertida ▼"
        elif curve_normalizing: curve_txt = "Normalizando ↗"
        elif curve_steep:       curve_txt = "Empinada ▲"
        elif curve_normal:      curve_txt = "Normal →"
        else:                   curve_txt = "Plana →"
    else:
        curve_txt = "N/A"

    y10_rising  = y10_chg is not None and y10_chg > 0
    y10_falling = y10_chg is not None and y10_chg < 0
    y10_txt = ("▲ Subiendo" if y10_rising else ("▼ Bajando" if y10_falling else "→ Estable")) if y10_val else "N/A"

    cli_expanding   = cli_val is not None and cli_val > 100.2
    cli_contracting = cli_val is not None and cli_val < 99.5
    cli_rising_bool = cli_chg is not None and cli_chg > 0
    cli_falling_bool= cli_chg is not None and cli_chg < 0

    if cli_val is not None:
        cli_base = "▲ Expansión" if cli_expanding else ("▼ Contracción" if cli_contracting else "→ Neutral")
        cli_txt  = cli_base + (" ↑" if cli_rising_bool else " ↓")
    else:
        cli_txt = "N/A"

    cfnai_positive = cfnai_val is not None and cfnai_val > 0.1
    cfnai_recession= cfnai_val is not None and cfnai_val < -0.7
    cfnai_txt = ("▲ Sólido" if cfnai_positive else ("⚠ Recesión" if cfnai_recession else "→ Débil")) if cfnai_val is not None else "N/A"
    fed_txt = "▲ Subiendo" if rates_rising else ("▼ Bajando" if rates_falling else "→ Pausa")

    # ── VOTOS ─────────────────────────────────────────────────────────────────
    votes = []

    # GDP (peso 8)
    if gdp_val is not None:
        if gdp_val > 3.0:       votes.append((EXP_T, 8))
        elif gdp_val > 1.5:     votes.append((EXP_T, 5)); votes.append((EXP_L, 3))
        elif gdp_val > 0:       votes.append((EXP_L, 5)); votes.append((EXP_T, 3))
        elif gdp_val > -1.0:    votes.append((REC_T, 4)); votes.append((REC_L, 4))
        else:                   votes.append((REC_T, 6)); votes.append((REC_L, 2))

    # Desempleo (peso 7)
    if unemp_chg is not None:
        if unemp_chg < -0.1:                            votes.append((EXP_T, 7))
        elif unemp_chg < 0:                             votes.append((EXP_T, 4)); votes.append((EXP_L, 3))
        elif unemp_chg < 0.1:                           votes.append((EXP_L, 4)); votes.append((REC_T, 3))
        elif not unemp_high:                            votes.append((REC_T, 4)); votes.append((EXP_L, 3))
        elif unemp_high and rates_falling:              votes.append((REC_L, 5)); votes.append((REC_T, 2))
        else:                                           votes.append((REC_T, 4)); votes.append((REC_L, 3))

    # CPI (peso 6)
    if cpi_val is not None:
        if cpi_val > 5.0:       votes.append((REC_T, 6))
        elif cpi_val > 3.0:     votes.append((EXP_L, 4)); votes.append((REC_T, 2))
        elif cpi_val > 2.0:     votes.append((EXP_L, 3)); votes.append((EXP_T, 3))
        elif cpi_val > 1.0:     votes.append((EXP_T, 3)); votes.append((REC_L, 3))
        else:                   votes.append((REC_L, 5)); votes.append((EXP_T, 1))

    # Fed (peso 6)
    if fed_val is not None:
        if rates_rising:                                        votes.append((EXP_L, 6))
        elif rates_falling and not gdp_growing and unemp_high: votes.append((REC_L, 6))
        elif rates_falling and gdp_growing:                     votes.append((EXP_T, 4)); votes.append((REC_L, 2))
        elif rates_falling:                                     votes.append((REC_L, 4)); votes.append((REC_T, 2))
        elif rates_high:                                        votes.append((EXP_L, 4)); votes.append((REC_T, 2))
        else:                                                   votes.append((EXP_L, 4)); votes.append((EXP_T, 2))

    # Curva 10Y-2Y (peso 7)
    if spread_val is not None:
        if spread_val > 1.5:    votes.append((EXP_T, 7))
        elif spread_val > 0.8:  votes.append((EXP_T, 4)); votes.append((EXP_L, 3))
        elif spread_val > 0.3:  votes.append((EXP_L, 4)); votes.append((EXP_T, 3))
        elif spread_val > 0:    votes.append((EXP_L, 4)); votes.append((REC_T, 3))
        elif spread_val > -0.5: votes.append((REC_T, 5)); votes.append((EXP_L, 2))
        else:                   votes.append((REC_T, 5)); votes.append((REC_L, 2))

    # OECD CLI (peso 8)
    if cli_val is not None:
        if cli_val > 101.0:     votes.append((EXP_T, 8))
        elif cli_val > 100.2:   votes.append((EXP_T, 5)); votes.append((EXP_L, 3))
        elif cli_val > 99.5:    votes.append((EXP_L, 4)); votes.append((EXP_T, 4))
        elif cli_val > 98.5:    votes.append((REC_T, 4)); votes.append((EXP_L, 4))
        elif cli_val > 97.0:    votes.append((REC_T, 6)); votes.append((REC_L, 2))
        else:                   votes.append((REC_T, 5)); votes.append((REC_L, 3))

    # CFNAI (peso 5)
    if cfnai_val is not None:
        if cfnai_val > 0.5:     votes.append((EXP_T, 5))
        elif cfnai_val > 0.1:   votes.append((EXP_T, 3)); votes.append((EXP_L, 2))
        elif cfnai_val > -0.1:  votes.append((EXP_L, 3)); votes.append((REC_T, 2))
        elif cfnai_val > -0.7:  votes.append((REC_T, 3)); votes.append((REC_L, 2))
        else:                   votes.append((REC_T, 3)); votes.append((REC_L, 2))

    # VIX MA25/200 (peso 4)
    vix_ma25 = vix_ma200 = None
    try:
        vix_df = yf.download("^VIX", period="2y", auto_adjust=True, progress=False)
        if not vix_df.empty:
            vix_series = vix_df["Close"].dropna()
            if hasattr(vix_series, "columns"):
                vix_series = vix_series.iloc[:, 0]
            if len(vix_series) >= 200:
                vix_ma25  = float(vix_series.rolling(25).mean().iloc[-1])
                vix_ma200 = float(vix_series.rolling(200).mean().iloc[-1])
    except Exception:
        pass

    vix_deteriorating = vix_ma25 is not None and vix_ma200 is not None and vix_ma25 > vix_ma200
    vix_improving     = vix_ma25 is not None and vix_ma200 is not None and vix_ma25 < vix_ma200
    vix_fear          = vix_ma25 is not None and vix_ma25 > 25
    vix_panic         = vix_ma25 is not None and vix_ma25 > 35

    if vix_ma25 is not None and vix_ma200 is not None:
        vix_trend_txt = "↑ Deteriorando" if vix_deteriorating else "↓ Mejorando"
        signals_vix = (vix_trend_txt, vix_ma25, vix_ma25 - vix_ma200)
        if vix_panic:                               votes.append((REC_T, 4))
        elif vix_fear and vix_deteriorating:        votes.append((REC_T, 3)); votes.append((EXP_L, 1))
        elif vix_deteriorating:                     votes.append((EXP_L, 2)); votes.append((REC_T, 2))
        else:                                       votes.append((EXP_T, 3)); votes.append((EXP_L, 1))
    else:
        signals_vix = ("N/A", None, None)

    # ── Calcular grados ───────────────────────────────────────────────────────
    degrees = weighted_circular_mean(votes)

    if degrees < 90:    phase_idx = 0
    elif degrees < 180: phase_idx = 1
    elif degrees < 270: phase_idx = 2
    else:               phase_idx = 3

    # Score legacy para compatibilidad
    score = [0, 0, 0, 0]
    score[phase_idx] = 10
    center = phase_idx * 90 + 45
    dist = abs(degrees - center)
    if dist > 30:
        adjacent = (phase_idx + 1) % 4 if degrees > center else (phase_idx - 1) % 4
        score[adjacent] = 5

    # ── Signals dict ─────────────────────────────────────────────────────────
    signals = {
        "GDP QoQ":      ("▲ Creciendo" if gdp_growing else "▼ Cayendo",   gdp_val,    gdp_chg),
        "Desempleo":    ("▼ Bajando"   if unemp_falling else "▲ Subiendo", unemp_val,  unemp_chg),
        "CPI YoY":      ("▲ Alto"      if cpi_high else "✓ Moderado",      cpi_val,    cpi_chg),
        "Fed Funds":    (fed_txt,                                           fed_val,    fed_yoy),
        "Curva 10Y-2Y": (curve_txt,                                        spread_val, spread_chg),
        "10Y Yield":    (y10_txt,                                          y10_val,    y10_chg),
        "OECD CLI":     (cli_txt,                                          cli_val,    cli_chg),
        "CFNAI":        (cfnai_txt,                                        cfnai_val,  cfnai_chg),
        "VIX MA25/200": signals_vix,
    }

    print(f"   [DEBUG] degrees: {degrees:.1f}° | phase: {PHASE_NAMES[phase_idx]}")

    return phase_idx, signals, score, degrees


def get_ema200_weekly(ticker, prices_daily):
    """
    Calcula EMA200 semanal a partir de precios diarios.
    Resamplea a semanas y calcula EMA de 200 periodos.
    """
    try:
        if ticker not in prices_daily.columns:
            return None, None
        s = prices_daily[ticker].dropna()
        if len(s) < 10:
            return None, None
        weekly = s.resample("W").last().dropna()
        if len(weekly) < 10:
            return None, None
        ema200 = weekly.ewm(span=200, adjust=False).mean()
        price   = float(s.iloc[-1])
        ema_val = float(ema200.iloc[-1])
        above   = price > ema_val
        pct_vs  = (price - ema_val) / ema_val
        return above, pct_vs
    except Exception:
        return None, None


# Sensibilidad de cada sector al entorno de tipos
# (+) se beneficia de tipos altos/subiendo  (-) se perjudica  (=) neutro
SECTOR_RATE_SENSITIVITY = {
    "XLK":  ("-", "Valoraciones presionadas con tipos altos"),
    "XLY":  ("-", "Consumo apalancado, sensible a tipos"),
    "XLC":  ("-", "Alto múltiplo, penalizado por descuento"),
    "XLF":  ("+", "Margen de interés neto mejora con tipos altos"),
    "XLI":  ("=", "Neutro, depende más del ciclo que de tipos"),
    "XLB":  ("=", "Materiales: más sensible a demanda global"),
    "XLP":  ("=", "Defensivo, poco sensible a tipos"),
    "XLV":  ("=", "Defensivo, relativamente inmune a tipos"),
    "XLU":  ("-", "Alto dividendo, compite con bonos al subir tipos"),
    "XLE":  ("+", "Correlación positiva con inflación y tipos altos"),
    "XLRE": ("-", "Muy sensible: coste financiación y descuento de renta"),
}

def build_spi_data(prices_daily, macro_results):
    """Construye todos los datos necesarios para la pestaña SPI."""
    phase_idx, signals, score, degrees = detect_cycle_phase(macro_results, prices_daily)

    # Determinar entorno de tipos para la columna de sensibilidad
    def get_val(region, name):
        for r, indicators in macro_results.items():
            for n, unit, val, chg_last, chg_yoy, note in indicators:
                if n == name:
                    return val, chg_last
        return None, None

    y10_val, y10_chg = get_val("US YIELD CURVE", "US 10Y Yield")
    rates_rising = y10_chg is not None and y10_chg > 0

    sector_data = []
    for name, ticker, w_et, w_el, w_rt, w_rl in SPI_SECTORS:
        weights    = [w_et, w_el, w_rt, w_rl]
        rec_weight = weights[phase_idx]

        above_ema, pct_ema = get_ema200_weekly(ticker, prices_daily)

        price, r1m, r3m, ytd, r1y = None, None, None, None, None
        if ticker in prices_daily.columns:
            s     = prices_daily[ticker].dropna()
            price = float(s.iloc[-1]) if len(s) > 0 else None
            r1m   = calc_return(s, 21)
            r3m   = calc_return(s, 63)
            ytd   = ytd_return(s)
            r1y   = calc_return(s, 252)

        alerta = (above_ema is not None and not above_ema and rec_weight >= 0.10)

        # Sensibilidad tipos: favorable/desfavorable según entorno actual
        rate_sign, rate_note = SECTOR_RATE_SENSITIVITY.get(ticker, ("=", ""))
        if rates_rising:
            rate_signal = "favorable" if rate_sign == "+" else ("desfav." if rate_sign == "-" else "neutro")
            rate_color  = "27AE60"    if rate_sign == "+" else ("E74C3C"   if rate_sign == "-" else SECTION_BG)
        else:  # tipos bajando
            rate_signal = "favorable" if rate_sign == "-" else ("desfav." if rate_sign == "+" else "neutro")
            rate_color  = "27AE60"    if rate_sign == "-" else ("E74C3C"   if rate_sign == "+" else SECTION_BG)

        sector_data.append({
            "name":        name,
            "ticker":      ticker,
            "weights":     weights,
            "rec_weight":  rec_weight,
            "price":       price,
            "r1m":         r1m,
            "r3m":         r3m,
            "ytd":         ytd,
            "r1y":         r1y,
            "above_ema":   above_ema,
            "pct_ema":     pct_ema,
            "alerta":      alerta,
            "rate_signal": rate_signal,
            "rate_color":  rate_color,
            "rate_note":   rate_note,
        })

    return phase_idx, signals, score, sector_data, degrees

# ── SHEET 3: SPI ──────────────────────────────────────────────────────────────

def write_spi_sheet(ws, phase_idx, signals, score, sector_data, today_str, degrees=45):
    ws.sheet_view.showGridLines = False

    phase_name  = PHASE_NAMES[phase_idx]
    phase_color = PHASE_COLORS[phase_idx]

    # ── Título ──
    ws.merge_cells("A1:H1")
    ws["A1"] = "Sector Pulse Investing (SPI)"
    ws["A1"].font = fnt(bold=True, size=14); ws["A1"].fill = fill(DARK_BG); ws["A1"].alignment = left(1)
    ws.merge_cells("I1:M1")
    ws["I1"] = today_str
    ws["I1"].font = fnt(bold=True, color=LIGHT_GRAY, size=11); ws["I1"].fill = fill(DARK_BG); ws["I1"].alignment = center()
    ws.row_dimensions[1].height = 22

    # ── Fase detectada ──
    ws.merge_cells("A2:M2")
    ws["A2"] = f"  FASE ACTUAL DEL CICLO:  {phase_name.upper()}"
    ws["A2"].font = Font(bold=True, color=WHITE, size=12, name="Arial")
    ws["A2"].fill = fill(phase_color)
    ws["A2"].alignment = left(1)
    ws.row_dimensions[2].height = 22

    # ── Indicador de grados ──
    ws.merge_cells("A3:M3")
    phase_center = phase_idx * 90 + 45
    dist_to_center = abs(degrees - phase_center)
    intensity = "central" if dist_to_center < 20 else ("límite" if dist_to_center > 35 else "moderada")
    deg_str = (f"  {degrees:.1f}°  |  "
               f"ExpT 0-90° · ExpL 90-180° · RecT 180-270° · RecL 270-360°  |  "
               f"Posición: {intensity} en {phase_name}")
    ws["A3"] = deg_str
    ws["A3"].font = fnt(italic=True, color=LIGHT_GRAY, size=8)
    ws["A3"].fill = fill(SECTION_BG); ws["A3"].alignment = left(1)
    ws.row_dimensions[3].height = 13

    # ── Cabecera bloque indicadores ──
    ws.merge_cells("A4:M4")
    ws["A4"] = "  6 Pilares del Ciclo  (PIB · Empleo · Inflación · Fed Funds · Curva de Tipos · 10Y Yield · VIX MA25/200)"
    ws["A4"].font = fnt(bold=True, color=LIGHT_GRAY, size=8)
    ws["A4"].fill = fill(HEADER_BG); ws["A4"].alignment = left(1)
    ws.row_dimensions[4].height = 13

    # ── Bloque de indicadores: 3 filas x 3 cols (A-I) + VIX en cols J-M ──────
    # Fila 5-7:  GDP · Desempleo · CPI
    # Fila 8-10: Fed Funds · Curva · 10Y
    # Fila 11-13: OECD CLI · CFNAI · VIX MA25/200
    positive_keywords = {"▲ Creciendo", "▼ Bajando", "✓ Moderado", "Empinada ▲",
                         "Normalizando ↗", "▲ Expansión ↑", "▲ Sólido"}

    def write_indicator_block(ws, row, col, ind_name, status, val, chg):
        c = ws.cell(row=row, column=col, value=ind_name)
        c.font = fnt(bold=True, color=MID_GRAY, size=8); c.fill = fill(DARK_BG); c.alignment = center()
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        val_str = f"{val:.2f}" if val is not None else "N/A"
        chg_str = f" ({'+' if chg and chg > 0 else ''}{chg:.2f})" if chg is not None else ""
        c2 = ws.cell(row=row+1, column=col, value=f"{val_str}{chg_str}")
        c2.font = fnt(bold=True, size=9); c2.fill = fill(DARK_BG); c2.alignment = center()
        ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
        is_green = any(k in status for k in ("Creciendo","Bajando","Moderado","Empinada","Normalizando","Expansión ↑","Sólido"))
        is_red   = any(k in status for k in ("Cayendo","Subiendo","Alto","Invertida","Contracción","Recesión","Deteriorando"))
        tc = "27AE60" if is_green else ("E74C3C" if is_red else MID_GRAY)
        c3 = ws.cell(row=row+2, column=col, value=status)
        c3.font = fnt(bold=True, color=tc, size=8); c3.fill = fill(DARK_BG); c3.alignment = center()
        ws.merge_cells(start_row=row+2, start_column=col, end_row=row+2, end_column=col+1)
        ws.row_dimensions[row].height   = 13
        ws.row_dimensions[row+1].height = 16
        ws.row_dimensions[row+2].height = 13

    # Fila 1 de indicadores (fila 5): GDP · Desempleo · CPI
    write_indicator_block(ws, 5, 1,  "GDP QoQ",   *signals["GDP QoQ"])
    write_indicator_block(ws, 5, 3,  "Desempleo", *signals["Desempleo"])
    write_indicator_block(ws, 5, 5,  "CPI YoY",   *signals["CPI YoY"])
    # Separador
    for r in range(5,8):
        for c in [7]: ws.cell(row=r, column=c).fill = fill(DARK_BG)

    # Fila 2 de indicadores (fila 8): Fed Funds · Curva · 10Y
    write_indicator_block(ws, 8, 1,  "Fed Funds",    *signals["Fed Funds"])
    write_indicator_block(ws, 8, 3,  "Curva 10Y-2Y", *signals["Curva 10Y-2Y"])
    write_indicator_block(ws, 8, 5,  "10Y Yield",    *signals["10Y Yield"])
    # Separador
    for r in range(8,11):
        for c in [7]: ws.cell(row=r, column=c).fill = fill(DARK_BG)

    # Fila 3 de indicadores (fila 11): OECD CLI · CFNAI · VIX MA25/200
    write_indicator_block(ws, 11, 1, "OECD CLI", *signals.get("OECD CLI", ("N/A", None, None)))
    write_indicator_block(ws, 11, 3, "CFNAI",    *signals.get("CFNAI",    ("N/A", None, None)))
    vix_sig = signals.get("VIX MA25/200", ("N/A", None, None))
    write_indicator_block(ws, 11, 5, "VIX MA25/200", *vix_sig)
    # Separador
    for r in range(11,14):
        for c in [7]: ws.cell(row=r, column=c).fill = fill(DARK_BG)

    # Columnas 8-13 con fondo oscuro en filas de indicadores
    for r in range(5, 14):
        for c in range(8, 14):
            ws.cell(row=r, column=c).fill = fill(DARK_BG)

    # ── Separador tras indicadores ──
    ws.row_dimensions[14].height = 6
    for c in range(1, 14):
        ws.cell(row=14, column=c).fill = fill(DARK_BG)

    # ── Sub-cabecera pesos ──
    ws.merge_cells("D15:G15")
    c = ws.cell(row=15, column=4, value="Pesos por fase")
    c.font = fnt(bold=True, color=LIGHT_GRAY, size=8)
    c.fill = fill(SECTION_BG); c.alignment = center()
    ws.row_dimensions[15].height = 13

    # ── Cabeceras tabla ──
    headers = ["Sector", "Ticker", "Precio",
               "Exp. Temp.", "Exp. Tard.", "Rec. Temp.", "Rec. Tard.",
               "Peso Actual", "1M", "3M", "YTD", "vs EMA200W", "Tipos"]
    widths  = [18, 7, 9, 10, 10, 10, 10, 11, 8, 8, 8, 12, 10]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=16, column=ci, value=h)
        c.font = fnt(bold=True, color=LIGHT_GRAY, size=8)
        c.fill = fill(HEADER_BG); c.alignment = center()
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[16].height = 15

    # ── Filas de sectores ──
    er = 17
    for sd in sector_data:
        alerta    = sd["alerta"]
        rec_w     = sd["rec_weight"]
        above_ema = sd["above_ema"]
        pct_ema   = sd["pct_ema"]
        row_bg    = "1F3244" if alerta else DARK_BG

        # Col 1: Nombre
        c = ws.cell(row=er, column=1, value=f"⚠ {sd['name']}" if alerta else sd["name"])
        c.font = fnt(bold=alerta, color="FFD700" if alerta else LIGHT_GRAY, size=8)
        c.fill = fill(row_bg); c.alignment = left(1)

        # Col 2: Ticker
        c = ws.cell(row=er, column=2, value=sd["ticker"])
        c.font = fnt(color=MID_GRAY, size=8, italic=True)
        c.fill = fill(row_bg); c.alignment = center()

        # Col 3: Precio
        c = ws.cell(row=er, column=3, value=fmt_price(sd["price"]))
        c.font = fnt(size=8); c.fill = fill(row_bg); c.alignment = center()

        # Cols 4-7: Pesos por fase
        for fi, w in enumerate(sd["weights"]):
            is_active = (fi == phase_idx)
            c = ws.cell(row=er, column=fi+4, value=f"{w*100:.0f}%")
            c.font = fnt(bold=is_active, size=8)
            c.fill = fill(phase_color if is_active else SECTION_BG); c.alignment = center()

        # Col 8: Peso actual destacado
        c = ws.cell(row=er, column=8, value=f"{rec_w*100:.0f}%")
        c.font = Font(bold=True, color=WHITE, size=10, name="Arial")
        c.fill = fill("1A6B3A" if rec_w >= 0.15 else SECTION_BG); c.alignment = center()

        # Cols 9-11: Retornos
        for ri, rv in enumerate([sd["r1m"], sd["r3m"], sd["ytd"]]):
            c = ws.cell(row=er, column=9+ri, value=fmt_pct(rv))
            c.font = fnt(color=text_ret(rv), size=8, bold=(rv is not None and abs(rv) > 0.05))
            c.fill = fill(color_ret(rv)); c.alignment = center()

        # Col 12: vs EMA200W
        if above_ema is None:
            ema_txt, ema_bg, ema_tc = "N/A", SECTION_BG, MID_GRAY
        elif above_ema:
            ema_txt, ema_bg, ema_tc = f"▲ +{pct_ema*100:.1f}%", "1A5C2A", "A9DFBF"
        else:
            ema_txt, ema_bg, ema_tc = f"▼ {pct_ema*100:.1f}%", "7B241C", "F1948A"
        c = ws.cell(row=er, column=12, value=ema_txt)
        c.font = fnt(bold=True, color=ema_tc, size=8); c.fill = fill(ema_bg); c.alignment = center()

        # Col 13: Sensibilidad tipos
        c = ws.cell(row=er, column=13, value=sd["rate_signal"])
        c.font = fnt(bold=True, color=WHITE, size=8)
        c.fill = fill(sd["rate_color"]); c.alignment = center()

        ws.row_dimensions[er].height = 15
        er += 1

    # ── Leyenda ──
    er += 1
    ws.merge_cells(f"A{er}:M{er}")
    ws.cell(row=er, column=1,
        value="  ⚠ Alerta: sector con peso ≥10% por debajo de EMA200W  |  Tipos: favorable/desfav. según entorno actual del 10Y  |  Score = suma de señales por fase").font = fnt(italic=True, color=MID_GRAY, size=7)
    ws.cell(row=er, column=1).fill = fill(DARK_BG)
    ws.cell(row=er, column=1).alignment = left(1)
    ws.row_dimensions[er].height = 13

    # ── Tabla referencia fases ──
    er += 2
    ws.merge_cells(f"A{er}:M{er}")
    c = ws.cell(row=er, column=1, value="  REFERENCIA: Sectores favorecidos y lógica por fase")
    c.font = fnt(bold=True, color=LIGHT_GRAY, size=9)
    c.fill = fill(SECTION_BG); c.alignment = left(1)
    ws.row_dimensions[er].height = 14; er += 1

    ref_data = [
        ("Expansión Temprana", "27AE60", "XLK (25%), XLY (20%), XLC (15%)  →  Tipos bajos/bajando, consumidor reactiva, tech lidera rebote"),
        ("Expansión Tardía",   "F39C12", "XLF (20%), XLI (20%), XLB (15%)  →  Tipos altos, curva plana/invertida, inflación en pico"),
        ("Recesión Temprana",  "E74C3C", "XLV (25%), XLP (20%), XLU (20%)  →  Curva invertida, tipos aún altos, defensivos como refugio"),
        ("Recesión Tardía",    "8E44AD", "XLE (15%), XLV (15%), XLRE (12%) →  Curva normalizando, tipos bajando, XLRE y XLU se reactivan"),
    ]
    for phase, color, desc in ref_data:
        c1 = ws.cell(row=er, column=1, value=phase)
        c1.font = fnt(bold=True, size=8); c1.fill = fill(color); c1.alignment = left(1)
        ws.merge_cells(start_row=er, start_column=1, end_row=er, end_column=3)
        c2 = ws.cell(row=er, column=4, value=desc)
        c2.font = fnt(color=LIGHT_GRAY, size=8); c2.fill = fill(DARK_BG); c2.alignment = left(1)
        ws.merge_cells(start_row=er, start_column=4, end_row=er, end_column=13)
        ws.row_dimensions[er].height = 14; er += 1

    ws.freeze_panes = "A11"

# ── DISCORD ───────────────────────────────────────────────────────────────────

PHASE_EMOJIS = ["🟢", "🟠", "🔴", "🟣"]

def build_highlights(market_rows, sector_data, signals, phase_idx):
    """
    Detecta eventos destacados del día:
    - Movimientos extremos 1D (>2%) en cualquier activo
    - Alertas SPI: sector con peso alto bajo EMA200W
    - Macro fuera de rango: VIX>25, curva cruzando 0
    """
    highlights = []

    # 1. Movimientos extremos en mercado (1D > ±2%)
    THRESHOLD_1D = 0.02
    for row in market_rows:
        if row[0] == "HEADER":
            continue
        name, ticker, price, _, _, r1d = row[0], row[1], row[2], row[3], row[4], row[5]
        if r1d is not None and abs(r1d) >= THRESHOLD_1D:
            emoji = "🟢" if r1d > 0 else "🔴"
            sign  = "+" if r1d > 0 else ""
            highlights.append(f"{emoji} **{name}** ({ticker})  {sign}{r1d*100:.1f}% en 1D")

    # 2. Alertas SPI: sector con peso ≥10% bajo EMA200W
    for sd in sector_data:
        if sd["alerta"]:
            highlights.append(
                f"⚠️ **{sd['name']}** ({sd['ticker']}) bajo EMA200W "
                f"({sd['pct_ema']*100:.1f}%) — peso actual {sd['rec_weight']*100:.0f}%"
            )

    # 3. Macro fuera de rango
    def get_signal(name):
        return signals.get(name, (None, None, None))

    _, vix_val, _      = get_signal("VIX") if "VIX" in signals else (None, None, None)
    _, spread_val, _   = get_signal("Curva 10Y-2Y")
    _, spread_chg, _   = (None, None, None)  # chg está en posición 2
    sig_curve          = signals.get("Curva 10Y-2Y", (None, None, None))
    spread_val         = sig_curve[1]
    spread_chg         = sig_curve[2]

    if spread_val is not None and spread_chg is not None:
        # Curva cruzando 0 (de negativo a positivo o viceversa)
        prev_spread = spread_val - spread_chg
        if (prev_spread < 0 and spread_val >= 0):
            highlights.append("📐 **Curva 10Y-2Y** cruzó 0 al alza — señal de normalización")
        elif (prev_spread >= 0 and spread_val < 0):
            highlights.append("📐 **Curva 10Y-2Y** cruzó 0 a la baja — señal de inversión")

    return highlights


def send_discord(webhook_url, output_file, phase_idx, score, signals, sector_data, market_rows, today_str, degrees=135):
    """Envía resumen + Excel adjunto al canal de Discord via webhook."""
    import os, json
    import requests

    if not webhook_url:
        print("⚠️  DISCORD_WEBHOOK_URL no configurada, saltando envío.")
        return

    print(f"   URL recibida: {webhook_url[:40]}... (longitud: {len(webhook_url)})")

    phase_name  = PHASE_NAMES[phase_idx]
    phase_emoji = PHASE_EMOJIS[phase_idx]

    # Grados compacto
    phase_center = phase_idx * 90 + 45
    dist_to_center = abs(degrees - phase_center)
    intensity = "central" if dist_to_center < 20 else ("límite" if dist_to_center > 35 else "moderada")
    score_str = f"**{degrees:.1f}°** — posición {intensity} en {PHASE_NAMES[phase_idx]}"

    # Top 3 sectores por peso en fase actual
    top3 = sorted(sector_data, key=lambda x: x["rec_weight"], reverse=True)[:3]
    sectors_str = "\n".join(
        f"  `{sd['ticker']}`  {sd['rec_weight']*100:.0f}%"
        f"  |  {('+' if sd['r1m'] and sd['r1m']>0 else '')}{sd['r1m']*100:.1f}% 1M"
        if sd['r1m'] is not None else f"  `{sd['ticker']}`  {sd['rec_weight']*100:.0f}%"
        for sd in top3
    )

    # Macro resumen
    def sig(name):
        s = signals.get(name, (None, None, None))
        return f"{s[1]:.2f}" if s[1] is not None else "N/A"

    macro_str = (
        f"GDP `{sig('GDP QoQ')}%`  ·  CPI `{sig('CPI YoY')}%`  ·  "
        f"Fed `{sig('Fed Funds')}%`  ·  Curva `{sig('Curva 10Y-2Y')}`  ·  10Y `{sig('10Y Yield')}%`"
    )

    # Destacados
    highlights = build_highlights(market_rows, sector_data, signals, phase_idx)
    if highlights:
        hl_str = "\n".join(f"  {h}" for h in highlights[:6])
    else:
        hl_str = "  Sin movimientos destacados hoy."

    # Construir embed
    embed = {
        "title": f"📊 Market Tracker  —  {today_str}",
        "color": int(PHASE_COLORS[phase_idx], 16),
        "fields": [
            {
                "name": f"{phase_emoji} FASE DEL CICLO",
                "value": f"**{phase_name.upper()}**\n{score_str}",
                "inline": False
            },
            {
                "name": "⚡ Destacado hoy",
                "value": hl_str,
                "inline": False
            },
            {
                "name": f"📈 Sectores recomendados ({phase_name})",
                "value": sectors_str,
                "inline": False
            },
            {
                "name": "🌍 Macro",
                "value": macro_str,
                "inline": False
            },
        ],
        "footer": {"text": "SPI · Sector Pulse Investing  |  Datos: Yahoo Finance + FRED"},
    }

    # Enviar embed
    try:
        r = requests.post(
            webhook_url,
            json={"embeds": [embed]},
            timeout=15
        )
        print(f"   Discord embed: {r.status_code} {r.text[:200]}")
    except Exception as e:
        print(f"   ⚠️  Error enviando embed Discord: {e}")

    # Enviar Excel como archivo adjunto
    try:
        filename = os.path.basename(output_file)
        with open(output_file, "rb") as f:
            r2 = requests.post(
                webhook_url,
                files={"file": (filename, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                timeout=30
            )
        print(f"   Discord Excel: {r2.status_code} {r2.text[:200]}")
    except Exception as e:
        print(f"   ⚠️  Error enviando Excel Discord: {e}")


def main():
    today_str = datetime.today().strftime("%d-%b-%y")
    output    = f"market_tracker_{datetime.today().strftime('%Y%m%d')}.xlsx"

    print("📡 Descargando precios de mercado...")
    all_tickers = [t for _, assets in SECTIONS for _, t in assets]
    # Añadir tickers SPI por si no están en SECTIONS
    spi_tickers = [t for _, t, *_ in SPI_SECTORS]
    all_tickers = list(set(all_tickers + spi_tickers))
    prices = get_prices(all_tickers)

    print("🔢 Calculando retornos...")
    market_rows = build_market_rows(prices)

    print("🌍 Descargando datos macro...")
    macro_data = build_macro_data()

    print("🔄 Calculando fase del ciclo SPI...")
    phase_idx, signals, score, sector_data, degrees = build_spi_data(prices, macro_data)
    print(f"   → Fase detectada: {PHASE_NAMES[phase_idx]}")
    save_spi_history(today_str, degrees, phase_idx, signals)

    print("📝 Generando Excel...")
    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Markets"
    write_market_sheet(ws1, market_rows, today_str)
    add_banner_to_sheet(ws1, "banner.png", ws1.max_row)

    ws2 = wb.create_sheet("Macro")
    write_macro_sheet(ws2, macro_data, today_str)
    add_banner_to_sheet(ws2, "banner.png", ws2.max_row)

    ws3 = wb.create_sheet("SPI")
    write_spi_sheet(ws3, phase_idx, signals, score, sector_data, today_str, degrees)
    add_banner_to_sheet(ws3, "banner.png", ws3.max_row)

    wb.save(output)
    print(f"\n✅ Guardado: {output}")
    print(f"   Pestaña 1 → Markets  |  Pestaña 2 → Macro  |  Pestaña 3 → SPI ({PHASE_NAMES[phase_idx]})")

    print("\n📣 Enviando a Discord...")
    import os
    webhook_url = os.environ.get("DISCORD_WEBHOOK_URL", "")
    send_discord(webhook_url, output, phase_idx, score, signals, sector_data, market_rows, today_str, degrees)

    print("\n📰 Generando newsletter Substack...")
    anthropic_key = os.environ.get("ANTHROPIC_API_KEY", "")
    substack_sid  = os.environ.get("SUBSTACK_SID", "")
    date_str      = datetime.today().strftime("%Y-%m-%d")

    # 1+2. Narrativa con Claude (busca el briefing del día via web search)
    narrative = None
    if anthropic_key:
        print("   🤖 Generando narrativa con Claude...")
        narrative = generate_narrative(
            anthropic_key, None,
            PHASE_NAMES[phase_idx], signals, sector_data, today_str, degrees, phase_idx
        )
        if narrative:
            print(f"   ✅ Narrativa generada ({len(narrative)} chars)")
    else:
        print("   ⚠️  ANTHROPIC_API_KEY no configurada")

    # 3. Enviar a Make → Make publica en Substack
    make_webhook = os.environ.get("MAKE_WEBHOOK_URL", "")
    if make_webhook:
        title     = f"Market Tracker {today_str} · {PHASE_NAMES[phase_idx]}"
        body_html = build_substack_html(
            narrative, phase_idx, PHASE_NAMES[phase_idx],
            signals, sector_data, score, today_str, degrees
        )
        print("   📤 Enviando a Make webhook...")
        send_to_make(make_webhook, title, body_html, today_str, PHASE_NAMES[phase_idx])
    else:
        print("   ⚠️  MAKE_WEBHOOK_URL no configurada")


# ── ADVFN + CLAUDE NARRATIVA + SUBSTACK ───────────────────────────────────────

def send_to_make(webhook_url, title, body_html, today_str, phase_name):
    """Envía los datos del newsletter a Make via webhook."""
    import urllib.request, json
    try:
        payload = json.dumps({
            "title":      title,
            "body_html":  body_html,
            "subtitle":   f"Fase del ciclo · Sectores · Macro · {today_str}",
            "date":       today_str,
            "phase":      phase_name,
        }).encode("utf-8")
        req = urllib.request.Request(
            webhook_url,
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=15) as r:
            resp = r.read().decode()
            print(f"   ✅ Make webhook disparado: {resp}")
            return True
    except Exception as e:
        print(f"   ⚠️  Error enviando a Make: {e}")
        return False


def scrape_advfn(date_str):
    """
    Obtiene el briefing diario de ADVFN.
    Devuelve el texto limpio o None si no está disponible.
    """
    import urllib.request, re

    url = f"https://www.advfn.com/world-daily-market-briefing/{date_str}"
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
            "Connection": "keep-alive",
        })
        with urllib.request.urlopen(req, timeout=15) as r:
            html = r.read().decode("utf-8", errors="ignore")

        # Eliminar scripts, styles, nav, header, footer
        html = re.sub(r'<(script|style|nav|header|footer|aside)[^>]*>.*?</\1>', ' ', html, flags=re.DOTALL|re.IGNORECASE)

        # Buscar el contenido principal del artículo — ADVFN usa div con clase "article-body" o similar
        # Intentamos varias estrategias en orden
        text = ""

        # Estrategia 1: buscar párrafos dentro de article o div.article-body
        patterns = [
            r'<article[^>]*>(.*?)</article>',
            r'<div[^>]*class=["\'][^"\']*article[^"\']*["\'][^>]*>(.*?)</div>',
            r'<div[^>]*class=["\'][^"\']*content[^"\']*["\'][^>]*>(.*?)</div>',
            r'<div[^>]*id=["\'][^"\']*content[^"\']*["\'][^>]*>(.*?)</div>',
        ]
        for pattern in patterns:
            match = re.search(pattern, html, re.DOTALL|re.IGNORECASE)
            if match:
                block = match.group(1)
                # Extraer texto de párrafos
                paragraphs = re.findall(r'<p[^>]*>(.*?)</p>', block, re.DOTALL|re.IGNORECASE)
                if paragraphs:
                    text = " ".join(re.sub(r'<[^>]+>', ' ', p).strip() for p in paragraphs)
                    text = re.sub(r'\s+', ' ', text).strip()
                    if len(text) > 300:
                        break

        # Estrategia 2: extraer todos los <p> del HTML si no encontramos artículo
        if len(text) < 300:
            paragraphs = re.findall(r'<p[^>]*>(.*?)</p>', html, re.DOTALL|re.IGNORECASE)
            # Filtrar párrafos cortos (menú, pie de página, etc.)
            good_paragraphs = []
            for p in paragraphs:
                clean = re.sub(r'<[^>]+>', ' ', p).strip()
                clean = re.sub(r'\s+', ' ', clean)
                if len(clean) > 80:  # solo párrafos sustanciales
                    good_paragraphs.append(clean)
            text = " ".join(good_paragraphs)

        text = re.sub(r'\s+', ' ', text).strip()[:3000]
        print(f"   🔍 ADVFN preview: {text[:300]!r}")
        return text if len(text) > 200 else None

    except Exception as e:
        print(f"   ⚠️  ADVFN no disponible: {e}")
        return None


def generate_narrative(api_key, advfn_text, phase_name, signals, sector_data, today_str, degrees=135, phase_idx=1):
    """
    Usa Claude API para generar la narrativa del newsletter.
    """
    import urllib.request, json

    def sig(name):
        s = signals.get(name, (None, None, None))
        return f"{s[1]:.2f}" if s[1] is not None else "N/A"

    gauge_svg = generate_gauge_svg(degrees, phase_name, phase_idx)
    phase_center = phase_idx * 90 + 45
    dist_to_center = abs(degrees - phase_center)
    intensity = "central" if dist_to_center < 20 else ("límite" if dist_to_center > 35 else "moderada")
    score_str = f"posición {intensity}"

    top3 = sorted(sector_data, key=lambda x: x["rec_weight"], reverse=True)[:3]
    top3_str = ", ".join(f"{sd['ticker']} ({sd['rec_weight']*100:.0f}%)" for sd in top3)

    macro_context = (
        f"GDP: {sig('GDP QoQ')}% | CPI: {sig('CPI YoY')}% | "
        f"Fed Funds: {sig('Fed Funds')}% | Curva 10Y-2Y: {sig('Curva 10Y-2Y')} | "
        f"10Y Yield: {sig('10Y Yield')}%"
    )

    advfn_section = (
        f"Busca en internet el resumen de mercados del día {today_str} usando estas fuentes en orden de preferencia:\n"
        f"1. ADVFN: https://www.advfn.com/world-daily-market-briefing/{today_str}\n"
        f"2. Edward Jones Daily Market Recap: https://www.edwardjones.com/us-en/market-news-insights/stock-market-news/daily-market-recap\n"
        f"3. Reuters, Bloomberg o MarketWatch si las anteriores no están disponibles.\n"
        f"Usa el contenido encontrado para contextualizar la narrativa con los eventos reales del día."
    )

    system_prompt = """Eres el autor de un newsletter de inversión en español llamado 'Market Tracker'.
Tu estilo es profesional pero cercano, directo y sin jerga innecesaria.
Escribes para inversores particulares con conocimientos intermedios-avanzados.
No usas bullet points ni listas — escribes en prosa fluida.
Nunca das recomendaciones de compra/venta explícitas — das contexto y perspectiva."""

    user_prompt = f"""Fecha: {today_str}

DATOS DEL TRACKER:
- Fase del ciclo económico detectada: {phase_name}
- Sectores con mayor peso en esta fase: {top3_str}
- Indicadores macro: {macro_context}

{advfn_section}

Escribe la introducción del newsletter diario. Debe:
1. Abrir con una frase que capture el estado del mercado hoy (usa el briefing ADVFN si está disponible)
2. Conectar el contexto de mercado con la fase del ciclo detectada ({phase_name})
3. Mencionar brevemente qué implica esto para los sectores recomendados
4. Cerrar con una frase que invite al lector a revisar el análisis completo y el Excel adjunto

Longitud: 3-4 párrafos. Tono: analítico, sin alarmismo, con perspectiva de medio plazo."""

    try:
        payload = json.dumps({
            "model": "claude-sonnet-4-6",
            "max_tokens": 2000,
            "system": system_prompt,
            "tools": [{"type": "web_search_20250305", "name": "web_search"}],
            "messages": [{"role": "user", "content": user_prompt}]
        }).encode("utf-8")

        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={
                "Content-Type": "application/json",
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01"
            },
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=60) as r:
            data = json.loads(r.read().decode())
        # Extraer y concatenar todos los bloques de texto (ignorar tool_use)
        import re
        text_blocks = [
            block["text"] for block in data.get("content", [])
            if block.get("type") == "text" and block.get("text", "").strip()
        ]
        if not text_blocks:
            return None
        text = "\n\n".join(text_blocks)
        # Limpiar marcadores de citas web [1], [2] etc. y saltos de línea excesivos
        text = re.sub(r"\[\d+\]", "", text)
        text = re.sub(r"\n{3,}", "\n\n", text)
        text = re.sub(r" {2,}", " ", text)
        return text.strip()
    except Exception as e:
        try:
            print(f"   ⚠️  Error generando narrativa: {e.code} — {e.read().decode()}")
        except:
            print(f"   ⚠️  Error generando narrativa: {e}")
        return None


def publish_substack(substack_sid, publication_slug, title, body_html, today_str):
    """
    Publica un post en Substack usando la API no oficial.
    Solo para suscriptores de pago (paid).
    """
    import urllib.request, json

    base_url = "https://substack.com/api/v1"
    # Limpiar la cookie — puede venir con o sin el prefijo "s:"
    sid_clean = substack_sid.strip()
    headers  = {
        "Content-Type": "application/json",
        "Cookie": f"substack.sid={sid_clean}",
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
        "Referer": "https://substack.com/",
        "Origin": "https://substack.com",
    }

    # 1. Crear draft
    try:
        draft_payload = json.dumps({
            "draft_title":    title,
            "draft_body":     body_html,
            "draft_subtitle": f"Fase del ciclo · Sectores · Macro · {today_str}",
            "audience":       "everyone",   # gratuito para todos
            "type":           "newsletter",
        }).encode("utf-8")

        req = urllib.request.Request(
            f"{base_url}/drafts",
            data=draft_payload,
            headers=headers,
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=15) as r:
            draft = json.loads(r.read().decode())
            draft_id = draft.get("id")
            print(f"   Draft creado: {draft_id}")

        if not draft_id:
            print("   ⚠️  No se obtuvo draft_id")
            return False

        # 2. Publicar el draft
        pub_payload = json.dumps({
            "send":           True,
            "share_automatically": False,
        }).encode("utf-8")

        req2 = urllib.request.Request(
            f"{base_url}/drafts/{draft_id}/publish",
            data=pub_payload,
            headers=headers,
            method="POST"
        )
        with urllib.request.urlopen(req2, timeout=15) as r2:
            result = json.loads(r2.read().decode())
            print(f"   Substack publicado: {result.get('id', 'ok')}")
            return True

    except Exception as e:
        try:
            print(f"   ⚠️  Error publicando en Substack: {e.code} — {e.read().decode()[:300]}")
        except:
            print(f"   ⚠️  Error publicando en Substack: {e}")
        return False


def generate_gauge_svg(degrees, phase_name, phase_idx):
    """Genera gauge como HTML puro compatible con Gmail."""
    import math

    phase_colors = ["#27AE60", "#F39C12", "#E74C3C", "#8E44AD"]
    color = phase_colors[phase_idx]

    # Convertir grados del ciclo a porcentaje del arco (0-100%)
    # 0° ciclo = izquierda | 360° = derecha
    pct = (degrees / 360) * 100

    # Barra de progreso con 4 segmentos de color
    segments = [
        ("#27AE60", "ExpT 0-90°"),
        ("#F39C12", "ExpL 90-180°"),
        ("#E74C3C", "RecT 180-270°"),
        ("#8E44AD", "RecL 270-360°"),
    ]

    seg_html = ""
    for sc, _ in segments:
        seg_html += f'<td style="width:25%; background:{sc}; height:14px; opacity:0.7;"></td>\n'

    # Marcador de posición
    marker_left = max(2, min(96, pct))

    return f"""
<table width="300" cellpadding="0" cellspacing="0" style="border-collapse:collapse; background:#1C2634; border-radius:8px; padding:12px; font-family:Arial,sans-serif;">
  <tr>
    <td colspan="4" style="padding:8px 8px 4px; text-align:center; color:#ffffff; font-size:13px; font-weight:bold;">
      {degrees:.1f}° — <span style="color:{color};">{phase_name}</span>
    </td>
  </tr>
  <tr>
    <td colspan="4" style="padding:0 8px 4px;">
      <table width="100%" cellpadding="0" cellspacing="1" style="border-collapse:separate; border-radius:4px; overflow:hidden;">
        <tr>{seg_html}</tr>
      </table>
    </td>
  </tr>
  <tr>
    <td colspan="4" style="padding:0 8px 4px; position:relative; height:12px;">
      <table width="100%" cellpadding="0" cellspacing="0">
        <tr>
          <td width="{marker_left:.0f}%" style="text-align:right;">
            <span style="display:inline-block; width:0; height:0; border-left:5px solid transparent; border-right:5px solid transparent; border-bottom:8px solid white;"></span>
          </td>
          <td width="{100-marker_left:.0f}%"></td>
        </tr>
      </table>
    </td>
  </tr>
  <tr>
    <td style="padding:0 8px 8px; text-align:left; color:#27AE60; font-size:10px;">ExpT</td>
    <td style="padding:0 8px 8px; text-align:center; color:#F39C12; font-size:10px;">ExpL</td>
    <td style="padding:0 8px 8px; text-align:center; color:#E74C3C; font-size:10px;">RecT</td>
    <td style="padding:0 8px 8px; text-align:right; color:#8E44AD; font-size:10px;">RecL</td>
  </tr>
</table>"""


def build_substack_html(narrative, phase_idx, phase_name, signals, sector_data, score, today_str, degrees=45):
    """Construye el HTML del post de Substack."""

    def sig(name):
        s = signals.get(name, (None, None, None))
        return f"{s[1]:.2f}" if s[1] is not None else "N/A"

    # Debug temporal
    print(f"   [DEBUG] score: {score}")
    print(f"   [DEBUG] signals: { {k: v[1] for k,v in signals.items()} }")

    phase_color_hex = {
        0: "#27AE60", 1: "#F39C12", 2: "#E74C3C", 3: "#8E44AD"
    }[phase_idx]

    gauge_svg = generate_gauge_svg(degrees, phase_name, phase_idx)
    phase_center = phase_idx * 90 + 45
    dist_to_center = abs(degrees - phase_center)
    intensity = "central" if dist_to_center < 20 else ("límite" if dist_to_center > 35 else "moderada")

    top_sectors = sorted(sector_data, key=lambda x: x["rec_weight"], reverse=True)[:5]
    sectors_rows = "".join(
        f"<tr><td><b>{sd['name']}</b></td><td>{sd['ticker']}</td>"
        f"<td>{sd['rec_weight']*100:.0f}%</td>"
        f"<td>{'▲' if sd['r1m'] and sd['r1m']>0 else '▼'} {abs(sd['r1m']*100):.1f}%</td></tr>"
        if sd['r1m'] is not None else
        f"<tr><td><b>{sd['name']}</b></td><td>{sd['ticker']}</td><td>{sd['rec_weight']*100:.0f}%</td><td>–</td></tr>"
        for sd in top_sectors
    )

    score_str = f"posición {intensity}"

    # Convertir markdown básico a HTML
    def md_to_html(text):
        if not text:
            return ""
        import re
        # Negrita
        text = re.sub(r'\*\*(.*?)\*\*', r'<strong>\1</strong>', text)
        # Cursiva
        text = re.sub(r'\*(.*?)\*', r'<em>\1</em>', text)
        # Separadores
        text = re.sub(r'\n---+\n', '<hr>', text)
        # Párrafos
        paragraphs = [p.strip() for p in text.split("\n") if p.strip() and p.strip() != "---"]
        return "".join(f"<p>{p}</p>" for p in paragraphs)

    narrative_html = md_to_html(narrative) if narrative else ""

    return f"""
<div style="font-family: Georgia, serif; max-width: 680px; margin: 0 auto; color: #1a1a1a;">

  <div style="background:{phase_color_hex}; color:white; padding:16px 20px; border-radius:8px; margin-bottom:24px; display:flex; align-items:center; gap:20px;">
    <div style="flex:1;">
      <div style="font-size:13px; text-transform:uppercase; letter-spacing:1px; opacity:0.85;">Fase del Ciclo · {today_str}</div>
      <div style="font-size:24px; font-weight:bold; margin-top:4px;">{phase_name.upper()}</div>
      <div style="font-size:12px; margin-top:8px; opacity:0.8;">{degrees:.1f}° — {score_str}</div>
    </div>
    <div>{gauge_svg}</div>
  </div>

  {narrative_html}

  <h3 style="border-bottom:2px solid #eee; padding-bottom:8px;">📈 Sectores recomendados</h3>
  <table style="width:100%; border-collapse:collapse; font-size:14px;">
    <thead>
      <tr style="background:#f5f5f5;">
        <th style="padding:8px; text-align:left;">Sector</th>
        <th style="padding:8px;">Ticker</th>
        <th style="padding:8px;">Peso</th>
        <th style="padding:8px;">1M</th>
      </tr>
    </thead>
    <tbody>{sectors_rows}</tbody>
  </table>

  <h3 style="border-bottom:2px solid #eee; padding-bottom:8px; margin-top:24px;">🌍 Indicadores Macro</h3>
  <table style="width:100%; border-collapse:collapse; font-size:14px;">
    <tbody>
      <tr style="background:#f9f9f9;">
        <td style="padding:8px;"><b>GDP QoQ</b></td><td style="padding:8px;">{sig('GDP QoQ')}%</td>
        <td style="padding:8px;"><b>CPI YoY</b></td><td style="padding:8px;">{sig('CPI YoY')}%</td>
      </tr>
      <tr>
        <td style="padding:8px;"><b>Fed Funds</b></td><td style="padding:8px;">{sig('Fed Funds')}%</td>
        <td style="padding:8px;"><b>10Y Yield</b></td><td style="padding:8px;">{sig('10Y Yield')}%</td>
      </tr>
      <tr style="background:#f9f9f9;">
        <td style="padding:8px;"><b>Curva 10Y-2Y</b></td><td style="padding:8px;" colspan="3">{sig('Curva 10Y-2Y')} pts — {signals.get('Curva 10Y-2Y', ('N/A',))[0]}</td>
      </tr>
    </tbody>
  </table>

  <p style="margin-top:32px; font-size:12px; color:#888; border-top:1px solid #eee; padding-top:16px;">
    Market Tracker · SPI Sector Pulse Investing · Datos: Yahoo Finance + FRED<br>
    El Excel completo con todos los activos está disponible en el canal de Discord.
  </p>

</div>
"""

if __name__ == '__main__':
    main()
